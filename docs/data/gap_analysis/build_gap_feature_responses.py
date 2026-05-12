"""Annotate the data-scientist's Optum-RWD gap-feature workbook with codebase-aware responses.

Input: original workbook with 5 author-provided columns (E2I Field, Purpose,
Notes, Questions, Comments + a stray F7 weighting note for engagement_score).

Output: same workbook plus a new column ``G`` "Claude Code Analysis &
Recommendations" answering each row using verified references to the E2I
codebase (file:line) and Optum CDM. A header label is also added to column F so
the previously-orphan F7 weighting note is identifiable as "Additional
Comments".

Run:
    python docs/data/gap_analysis/build_gap_feature_responses.py
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path("/root/.claude/uploads/a3f58765-2c54-4bff-b35d-19e2f0cf0daa/8a7565c8-Questions_around_Gap_features.xlsx")
DST = Path(__file__).resolve().parent / "Questions_around_Gap_features_annotated.xlsx"

# Per-row response keyed by E2I Field. Each response is a multi-paragraph
# answer grounded in actual codebase references (file:line) and Optum CDM
# semantics, intended to be read in-cell by the data scientist.
RESPONSES: dict[str, str] = {
    "patient_journeys.journey_stage": (
        "PARAMETER VALIDATION:\n"
        "- observation_days=180: ALIGNS with scripts/convert_optum_rwd.py constants LOOKBACK_DAYS=180 and PREDICTION_DAYS=180. Use as-is.\n"
        "- pdc_threshold=0.80: INDUSTRY STANDARD (CMS Star Ratings, PQA). The 0.80 cutoff is also referenced in database/migrations/006_feedback_loop_infrastructure.sql. Adopt.\n"
        "- gap_threshold=60d: PARTIALLY OK. Our converter already differentiates by drug class: BIOLOGIC_DISCONT_GAP_DAYS=90 (Xolair/Dupixent are monthly injectables, so 60d is too aggressive and creates false discontinuations) and BIOLOGIC_PERSISTENCE_GAP_DAYS=60 (correct). Recommend keeping 60d for oral chronic therapies (anti-histamines, immunosuppressants), 90d for biologics. See convert_optum_rwd.py:90-91.\n\n"
        "SCHEMA REALITY CHECK:\n"
        "The 7-stage funnel (aware -> considering -> prescribed -> first_fill -> adherent -> discontinued -> maintained) does NOT exist in the schema today. journey_stage_type enum (database/core/e2i_ml_complete_v3_schema.sql:151) has only 5 values: diagnosis, initial_treatment, treatment_optimization, maintenance, treatment_switch. Until the enum is extended, derive but DOWN-MAP:\n"
        "- aware  -> diagnosis (dx code present, no Rx)\n"
        "- considering -> diagnosis (dx + specialist visit, still no Rx)\n"
        "- prescribed/first_fill -> initial_treatment\n"
        "- adherent (PDC>=0.80) -> maintenance\n"
        "- discontinued (gap exceeded) -> treatment_switch\n"
        "- maintained (>=180d on-therapy, no rescue events) -> maintenance\n\n"
        "DERIVATION RULES (Optum-derivable):\n"
        "Stage is per-patient + per-journey; recompute at each ETL run from the lookback window.\n"
        "- 'aware'      = dx code present (L50.x via demographics.diagcode or any inpatient diag1..5); no medication record for the indication.\n"
        "- 'considering'= 'aware' + >=1 procedure with allergist/derm taxonomy (provider.taxonomy1 prefix 207K/207N).\n"
        "- 'prescribed' = first appearance in medication.parquet for an in-class drug (CSU_BIOLOGIC_NDC_PREFIXES / CSU_BIOLOGIC_GENERICS).\n"
        "- 'first_fill' = same as 'prescribed' but with non-null days_sup and a paid claim (std_cost is not null OR copay+coins+deduct > 0).\n"
        "- 'adherent'   = rolling 180d PDC>=0.80 (sum days_sup / 180), no gap > gap_threshold.\n"
        "- 'discontinued'= last_fill_end_date + gap_threshold < today AND no subsequent fill.\n"
        "- 'maintained' = adherent for >=2 consecutive 180d windows, no rescue oral steroid burst, no urticaria/angioedema ED visits.\n\n"
        "ACTION: file a schema migration to extend journey_stage_type before populating; otherwise pipeline-side rules will silently down-map and lose granularity."
    ),

    "patient_journeys.risk_score": (
        "CURRENT STATE: scripts/convert_optum_rwd.py:1187 sets risk_score=None. Column is DECIMAL(3,2) in e2i_ml_complete_v3_schema.sql:853. The tool_composer/tool_registrations.py has a risk_scorer() stub returning toy values (0.82/0.45/0.12) but no real implementation.\n\n"
        "ANSWERS TO YOUR QUESTIONS:\n"
        "1) PURPOSE: For the CSU Optum cohort, build a THERAPY-RELATED RISK score (specifically discontinuation/non-adherence risk for biologic-treated patients). Rationale: it is the actionable use case for E2I sales-rep triggers and matches the existing targets (`discontinued_180d`, `persistent_at_180d` in convert_optum_rwd.py).\n\n"
        "2) TARGET: discontinued_180d (binary) is the most operationally useful single target. CATE on biologic discontinuation gives heterogeneous_effect estimates the heterogeneous_optimizer agent can act on.\n\n"
        "3) INPUTS: Use the existing §7 feature catalog already generated by _compute_features (convert_optum_rwd.py:736-960). DO NOT engineer new features for the first pass; the catalog is leakage-safe and well-documented:\n"
        "   - Demographics: age_at_index, gender, zip3, insurance_product, plan_type, urban_rural_code\n"
        "   - CSU disease: dx_l50_1/8/9_count, dx_angioedema_count, csu_chronicity\n"
        "   - Comorbidities (8 prefixes x 2 cols): has_<cond> + <cond>_claim_count + atopy_score + mental_health_flag + elixhauser_score + charlson_score\n"
        "   - Utilization: office_visits_total / _allergist / _dermatology, ed_visits_*, hospitalizations_total, unique_providers\n"
        "   - Drug exposure (7 classes x 4 cols): <class>_ever_filled, _fill_count, _days_supply_total, _days_since_last_fill\n"
        "   - Labs (8 LOINCs x 3 cols): <lab>_tested, _result_last, _abnormal_flag\n"
        "   - Provider mix: primary_specialist_type, saw_allergist_flag, saw_dermatologist_flag, specialist_concentration (HHI)\n\n"
        "4) COMORBIDITY SCORES: Charlson and Elixhauser are already computed (convert_optum_rwd.py:987-1022) but as APPROXIMATIONS — Elixhauser uses ICD-10 first-letter chapters; Charlson uses 5 high-severity categories (MI/CHF/cancer/diabetes/renal). For a research-grade model, swap with Quan or Sundararajan validated mappings (the docs/OPTUM_CONVERSION.md `Known approximations` table flags this as TODO).\n\n"
        "5) LOOKBACK: 180d (matches LOOKBACK_DAYS). Do NOT extend — the cohort's prediction window is also 180d post-index, so a longer lookback risks overlapping target observation.\n\n"
        "6) MODEL & OUTPUT: gradient-boosted classifier (xgboost or lightgbm — both already in `model_type` enum). Calibrate (Platt or isotonic) then write the probability to patient_journeys.risk_score (clamp 0.00-9.99 to fit DECIMAL(3,2); recommended convention: score = 10 * probability, so risk_score in [0.00, 9.99]). Persist SHAP values into ml_predictions.shap_values (the schema already supports this).\n\n"
        "DO NOT mix in disease-specific (e.g., 10y CVD or 5y diabetes) and therapy-related risk into one score; keep them as separate prediction_type values ('risk' vs 'churn' per the prediction_type enum in e2i_ml_v3_schema.sql)."
    ),

    "patient_journeys.data_quality_score": (
        "CURRENT IMPLEMENTATION: convert_optum_rwd.py:1160-1163 computes dq_score = non_null_features / total_features at the JOURNEY (patient) level — a uniform completeness ratio across all §7 features. Stored as DECIMAL(3,2). A separate aggregate-level scorer exists in database/ml/012_data_sources.sql (calculate_data_quality_score) with weights 30/30/20/20 for coverage/completeness/freshness/match_rate, but operates on source-level, not record-level.\n\n"
        "ANSWERS TO YOUR QUESTIONS:\n"
        "1) PER-CLAIM vs PER-PATIENT? Use HYBRID:\n"
        "   - Compute claim-level completeness for each row in medication/procedure/lab/inpatient\n"
        "   - Average per patient across the lookback window\n"
        "   - Persist patient-level score on patient_journeys.data_quality_score (matches current schema and existing converter behavior)\n"
        "   - Persist claim-level score on treatment_events as a new column (or in a side table) if granular drill-down is required\n\n"
        "2) COST FIELDS — RANKING & TREATMENT:\n"
        "   PRIMARY (most reliable in Optum CDM): std_cost\n"
        "   SECONDARY (patient OOP, billing-source dependent): charge, copay, coins, deduct\n"
        "   PHARMACY-ONLY (Rx claims only): dispfee, avgwhlsl\n"
        "   Rule: cost_completeness_flag = 1 if std_cost IS NOT NULL ELSE 0.5 if any(charge, copay, coins, deduct) IS NOT NULL ELSE 0. Treat dispfee/avgwhlsl as OPTIONAL — do not penalize medical claims for missing them.\n\n"
        "3) YOUR PROPOSED WEIGHTS (diagnosis 0.34 / procedure 0.33 / cost 0.33) — REASONABLE but suboptimal:\n"
        "   - Diagnosis fields drive cohort eligibility (without dx codes you can't even include the patient), so they deserve a heavier weight.\n"
        "   - Cost fields are 10-30% null in many Optum extracts (especially capitated MA plans), so down-weighting them avoids systematic penalty for plan-type effects.\n"
        "   RECOMMENDED WEIGHTS: diagnosis_weight=0.40, procedure_weight=0.25, cost_weight=0.20, enrollment_weight=0.15.\n"
        "   The ENROLLMENT dimension (eligeff/eligend non-null and continuous_enrollment==1) is currently treated as a hard exclusion in convert_optum_rwd.py — promote it to a soft quality dimension so partially-enrolled patients aren't silently dropped.\n\n"
        "4) FORMULA (per-claim, then averaged per patient):\n"
        "   claim_dqs = diagnosis_weight * dx_complete\n"
        "             + procedure_weight * proc_complete\n"
        "             + cost_weight      * cost_complete\n"
        "             + enrollment_weight* enroll_complete\n"
        "   patient_dqs = mean(claim_dqs over lookback window), then scale to DECIMAL(3,2) and store on patient_journeys.\n\n"
        "5) Hold up our existing data_source_tracking.data_quality_score (e2i_ml_v3_schema.sql:1395) for source-level monitoring — do NOT replace it; the record-level score is a finer-grained complement, not a substitute."
    ),

    "hcp_profiles.priority_tier": (
        "SCHEMA REALITY CHECK: priority_tier is INTEGER CHECK BETWEEN 1 AND 5 (e2i_ml_v3_schema.sql:788), where 1 = highest priority. Your proposed 3-tier mapping (High/Med/Low) must be reconciled with the 5-tier storage. Current Optum converter sets priority_tier=None (convert_optum_rwd.py:1445).\n\n"
        "RECOMMENDED 5->3 MAPPING (compatible with both schema and your proposal):\n"
        "   Tier 1 = decile 10                 (top 10%) — HIGH\n"
        "   Tier 2 = decile 8-9               (next 20%) — HIGH\n"
        "   Tier 3 = decile 4-7              (middle 40%) — MEDIUM\n"
        "   Tier 4 = decile 2-3            (next-lowest 20%) — LOW\n"
        "   Tier 5 = decile 1                (bottom 10%) — LOW\n"
        "Persist the 5-bin tier; surface 3-bin label as a presentation-layer view.\n\n"
        "DATA SOURCE & GRAIN:\n"
        "- ZIP_R counterpart in our codebase = medication.parquet (Rx fills with NPI). Lab equivalent (zip5_lr) is lab.parquet. The naming convention differs; the semantic content matches.\n"
        "- AGGREGATION KEY: (hcp_id, zip3, rolling_12mo). DO NOT use ZIP5 — too few HCPs per bin, decile is unstable. ZIP3 gives ~900 bins nationally with healthy HCP counts.\n"
        "- ROLLING WINDOW: 12 months ending at index_date. Avoids leakage into the prediction window.\n"
        "- DENOMINATOR: therapeutic-area TRx in the HCP's ZIP3 (CSU drugs only, not all Rx). Definition aligns with the NON_TARGET_DRUG_CLASSES + CSU_BIOLOGIC list in convert_optum_rwd.py:106-155.\n\n"
        "DECILE COMPUTATION:\n"
        "TRx count by HCP within ZIP3 -> rank -> assign decile via 10-bin equal-frequency. Ties broken by NDC-distinct count then alphabetical NPI. For HCPs with TRx=0 in window, assign Tier 5 (do NOT exclude — they are the 'no-treat' pool the model needs to score)."
    ),

    "hcp_profiles.adoption_category": (
        "BRAND SET (canonical, per convert_optum_rwd.py:106-112):\n"
        "   CSU_BIOLOGIC_BRANDS    = ('XOLAIR', 'DUPIXENT')\n"
        "   CSU_BIOLOGIC_GENERICS  = ('omalizumab', 'dupilumab')\n"
        "   CSU_BIOLOGIC_NDC_PREFIXES = ('50242', '00024', '0024')\n"
        "   CSU_BIOLOGIC_HCPCS    = {'J2357', 'J0517'}\n"
        "Use NDC prefix and HCPCS as canonical filters — brand strings in vendor data are inconsistent (XOLAIR vs Xolair vs XOLAIR (omalizumab)).\n\n"
        "WHY YOUR zip5_r DOES NOT CONTAIN 'Xolair':\n"
        "1. Xolair is often administered as a buy-and-bill biologic injectable -> the claim lands in PROCEDURE (HCPCS J2357), NOT pharmacy. Check procedure.parquet, not just medication/zip5_r.\n"
        "2. Vendor may have lemmatized brand to generic_name = 'omalizumab' -> filter on generic, not brand string.\n"
        "3. The NDC prefix 50242 reliably identifies Xolair regardless of brand string.\n"
        "Try in order: NDC prefix -> generic_name -> HCPCS -> brand_name fuzzy.\n\n"
        "BRAND LAUNCH DATES (FDA approval for CSU indication, anchor for time-to-first-Rx):\n"
        "- Xolair (omalizumab) for CSU: 2014-03-21 (already approved for asthma in 2003; CSU added Mar 2014). USE 2014-03-21 as launch.\n"
        "- Dupixent (dupilumab) for CSU: NOT FDA-approved for CSU as of 2026-05 (approved for AD/asthma/EoE/PN/COPD). For CSU adoption analysis treat Dupixent fills as OFF-LABEL or EXPANDED-ACCESS — do not anchor a CSU launch date for it; you can still compute adoption against the asthma launch (2018-10-19) if cross-indication adoption is of interest, but mark these HCPs separately.\n\n"
        "ROGERS DIFFUSION CURVE — USE CUMULATIVE-SHARE THRESHOLDS, NOT EQUAL QUARTILES:\n"
        "   Innovator       : first  2.5%   of adopters\n"
        "   Early adopter   : next  13.5%   (cum  16%)\n"
        "   Early majority  : next  34.0%   (cum  50%)\n"
        "   Late majority   : next  34.0%   (cum  84%)\n"
        "   Laggard         : last  16.0%   (cum 100%)\n"
        "WARNING: convert_optum_rwd.py:1412-1419 currently uses VOLUME quartiles (q25/q50/q75) and is MISSING the 'laggard' category — this is a known gap and should be re-implemented as time-to-first-Rx Rogers cutoffs.\n\n"
        "METHOD:\n"
        "for each HCP with >=1 fill of in-scope brand:\n"
        "    days_to_first = min(fill_dt) - brand_launch_date\n"
        "Rank HCPs by days_to_first ascending; assign Rogers category by cumulative share.\n"
        "HCPs with NO fill of the brand within the observation window: NOT 'laggard' — they are NON-ADOPTERS (separate category, do not pollute the diffusion curve)."
    ),

    "hcp_profiles.engagement_score": (
        "CORE FACT: Optum RWD contains NO Veeva CRM data. engagement_score MUST come from either (a) a separate Veeva integration, OR (b) Optum-derivable proxies. The current Optum converter sets digital_engagement_score=None (convert_optum_rwd.py:1453); feature_repo/features/hcp_features.py declares the FeatureView with ttl=1d, meaning the design EXPECTS a daily Veeva feed.\n\n"
        "YOUR PROPOSED WEIGHTING (F7 cell, recapped):\n"
        "  engagement_score = 0.35*visit_frequency + 0.20*activity_completion + 0.20*tactic_execution + 0.15*insight_engagement + 0.10*key_message\n"
        "This is REASONABLE for the Veeva pathway and aligns with the engagement_level cutoffs you proposed (>=75 High, >=40 Medium). Adopt for Tier 2 (Veeva-integrated) deliverable.\n\n"
        "VEEVA TABLE MAPPING (for the future integration):\n"
        "- visit_frequency_score   <- Veeva Call_vod__c (Activity)\n"
        "- activity_completion     <- Veeva Activity table (% of planned activities completed)\n"
        "- tactic_execution        <- Veeva Account_Tactic_vod__c -- YOUR PROPOSED 'Account Tactic' PROXY IS VALID as the substitute for trigger acceptance until E2I triggers ship\n"
        "- insight_engagement      <- Veeva Insight_vod__c or Survey_vod__c (positive/interest responses)\n"
        "- key_message_score       <- Veeva Key_Message_vod__c (delivery + receptivity)\n\n"
        "INTERIM OPTUM-ONLY PROXIES (claims-derivable, low-confidence):\n"
        "- prescribing_velocity = 90-day rolling TRx slope (proxy for receptiveness)\n"
        "- cross_indication_breadth = number of unique therapeutic areas Rx'd (proxy for openness to new molecules)\n"
        "- time_to_first_in_class = days from a new launch to the HCP's first Rx of any molecule in the class (proxy for innovation appetite)\n"
        "Combine these via PCA or simple z-score average; document explicitly as a CLAIMS-PROXY, not the Veeva engagement metric.\n\n"
        "STORAGE NOTES:\n"
        "- hcp_profiles.digital_engagement_score is DECIMAL(3,2) i.e. 0.00-9.99\n"
        "- Your 0-100 cutoffs (75/40) translate to 7.5 / 4.0 on the stored scale\n"
        "- patient_journeys.engagement_score is DECIMAL(4,2) with CHECK 0<=v<=10 — same scale, different table (per-patient HCP engagement at the journey index date)\n\n"
        "RECOMMENDATION: For Optum-only Tier 1 deliverable, leave engagement_score NULL with a documented gap in data_dictionary.csv. Adopt your weighting scheme when Veeva pipeline lands; do not block the Tier 1 model on Veeva availability."
    ),

    "triggers.*": (
        "YES — the triggers table already exists with full delivery/acceptance tracking. Reference: database/core/e2i_ml_complete_v3_schema.sql + data dictionary at docs/data/02-CORE-DATA-DICTIONARY.md:1068-1145.\n\n"
        "FIELDS COVERED (all your asks plus more):\n"
        "- trigger_id (VARCHAR(30) PK)\n"
        "- hcp_id (FK to hcp_profiles), patient_id\n"
        "- trigger_type (treatment_switch / adherence_risk / new_patient / ...)\n"
        "- delivery_channel (email / dashboard / push / crm)\n"
        "- delivery_status (pending / delivered / failed / expired) + delivery_timestamp\n"
        "- view_timestamp\n"
        "- acceptance_status (accepted / rejected / deferred / pending) + acceptance_timestamp\n"
        "- action_taken (text) + action_timestamp\n"
        "- false_positive_flag (BOOLEAN)\n"
        "- outcome_tracked + outcome_value\n"
        "- change tracking: previous_trigger_id, change_type, change_reason, change_failed, change_outcome_delta (these power the Change-Fail Rate KPI per docs/data/02-CORE-DATA-DICTIONARY.md:1131-1145)\n\n"
        "POPULATION REALITY:\n"
        "- Triggers are E2I-GENERATED, not from claims. No production TriggerGenerator agent in the codebase (search confirmed). Current rows are synthetic test data via src/ml/synthetic/generators/trigger_generator.py.\n"
        "- For the Optum cohort: leave the triggers table EMPTY during the data conversion. Triggers will be created downstream by Tier 2 agents (heterogeneous_optimizer + causal_impact) once the model is trained, and written via API.\n"
        "- Delivery tracking (delivery_*, acceptance_*, action_*) requires a Veeva CRM integration -> rep activity table -> triggers table sync. That pipeline is not yet built; do not attempt to populate from Optum alone.\n\n"
        "ROI MAPPING: src/agents/gap_analyzer/nodes/roi_calculator.py wires 'trigger_acceptance' -> ValueDriverType.ACTION_RATE (line ~60). Confirms the field is the authoritative one for KPI computation; your Veeva 'Account Tactic' proxy is correct as a transitional fill-in."
    ),

    "business_metrics.market_share": (
        "SCOPE — Xolair + Dupixent + COMPARATOR MARKET BASKET (NOT just the two biologics in isolation; share is a ratio):\n"
        "  Numerator      = TRx for the target brand (e.g., Xolair) in time t\n"
        "  Denominator    = TRx for the full CSU therapy basket in time t\n"
        "The CSU therapy basket is already defined in convert_optum_rwd.py:106-155 — adopt directly:\n"
        "- Biologics (numerator candidates): Xolair, Dupixent\n"
        "- Antihistamines (basket): cetirizine, loratadine, fexofenadine, desloratadine, levocetirizine (2g); diphenhydramine, hydroxyzine (1g)\n"
        "- H2 blockers: famotidine, ranitidine, cimetidine\n"
        "- LTRA: montelukast, zafirlukast\n"
        "- Steroids: prednisone, methylprednisolone, dexamethasone (systemic); triamcinolone, hydrocortisone, clobetasol (topical)\n"
        "- Immunosuppressants: cyclosporine, methotrexate, azathioprine, mycophenolate\n\n"
        "FORMULA:\n"
        "  market_share[brand, geography, t] = brand_TRx[brand, geography, t] / sum(basket_TRx[*, geography, t])\n"
        "Time grain: monthly (matches business_metrics.metric_date cadence in e2i_ml_v3_schema.sql).\n"
        "Geography grain: territory or ZIP3; aggregate to region/national.\n\n"
        "CRITICAL CAVEAT — OPTUM IS NOT IQVIA LAAD:\n"
        "Optum CDM is a single-payer panel (UnitedHealth-dominated, ~20% national market coverage), NOT all-payer like IQVIA LAAD. So market_share derived from Optum is a UHG-weighted approximation, not true national share. Document this in business_metrics.methodology_notes and set data_source = 'OPTUM_CDM' (not IQVIA_LAAD). For board-level reporting, IQVIA LAAD remains the authoritative source.\n\n"
        "MULTI-BRAND EXTENSION:\n"
        "Per brand_type enum (e2i_ml_v3_schema.sql), the supported brands include Remibrutinib (CSU), Fabhalta (PNH), Kisqali (HR+/HER2- breast cancer), competitor, other. Use the SAME formula but swap the indication-specific basket:\n"
        "- Remibrutinib: CSU basket (same as above, Remi joins biologics)\n"
        "- Fabhalta: PNH basket (eculizumab, ravulizumab, danicopan, pegcetacoplan)\n"
        "- Kisqali: HR+/HER2- breast cancer basket (palbociclib, abemaciclib, ribociclib)\n\n"
        "ANSWER TO YOUR QUESTION: For the current Optum-CSU work, START with Xolair + Dupixent only — the converter only loads CSU-relevant data anyway. Extend to other brands when their cohorts are built."
    ),

    "treatment_events.treatment_response": (
        "AGREEMENT WITH YOUR ANALYSIS: A single lab result cannot define treatment response — confirmed. The fields you listed (rslt_nbr, abnl_cd, loinc_cd, tst_desc, low_nrml, hi_nrml) provide per-event lab evidence only.\n\n"
        "CSU-SPECIFIC REALITY: CSU has NO validated lab biomarker for clinical control. Standard CSU response measures are PATIENT-REPORTED (UAS7, UCT, CU-Q2oL) — none of which are in Optum claims. Do not attempt a lab-based response derivation for CSU; use CLAIM-PATTERN PROXIES instead:\n\n"
        "CLAIM-PATTERN PROXY RULES (CSU cohort):\n"
        "Pre-conditions:\n"
        "  - Treatment initiated (>=1 fill of in-scope biologic)\n"
        "  - Persistence: >=60 consecutive days of biologic coverage by days_supply\n"
        "  - Follow-up: >=90 days post-initiation observation\n\n"
        "Classification:\n"
        "  - 'controlled' = persistence threshold met AND no rescue oral steroid burst (prednisone >=10mg >=5 days) AND no ED visit for urticaria (L50.x) or angioedema (T78.3) in the post-init window\n"
        "  - 'inadequate' = persistence met, but >=1 rescue steroid burst OR >=1 urticaria/angioedema ED visit\n"
        "  - 'refractory' = treatment switch to a second biologic OR addition of immunosuppressant (cyclosporine/methotrexate/azathioprine/mycophenolate per NON_TARGET_DRUG_CLASSES.immunosupp) within 180d\n"
        "  - 'discontinued' = gap > BIOLOGIC_DISCONT_GAP_DAYS=90d -- already computed by _target_discontinued_180d in convert_optum_rwd.py:1078\n\n"
        "EXISTING ENUM ALIGNMENT:\n"
        "- treatment_events.outcome_indicator (e2i_v3 schema) = improved / stable / worsened\n"
        "- treatment_response values used in src/kpi/calculators/brand_specific.py:123 = inadequate / uncontrolled / refractory\n"
        "Map the proxy categories above to these existing values to avoid schema churn.\n\n"
        "ADJACENT-AREA EXTENSION (NOT for CSU):\n"
        "For therapeutic areas WITH validated lab biomarkers, use longitudinal trajectories:\n"
        "- Fabhalta/PNH: LDH (LOINC 14804-9) baseline -> month-3 -> month-6 trajectory; response = LDH normalization (<=ULN) sustained.\n"
        "- Kisqali/HR+/HER2- BC: CA15-3, tumor markers (longitudinal); requires imaging/RECIST data not in claims.\n"
        "- Diabetes: HbA1c (LOINC 4548-4) trajectory; response = HbA1c reduction >=1.0% sustained 3 mo.\n"
        "Build these as LOINC-specific rules only when those cohorts are ramped; CSU stays claim-pattern-based.\n\n"
        "LOINC PANELS ALREADY CAPTURED for CSU (convert_optum_rwd.py:125-134): IgE total, eosinophil, CRP, TPO Ab, free-T4, TSH, ANA, CBC. These inform CSU ENDOTYPE (autoimmune vs idiopathic) and ENROLLMENT (autoimmune subtype responds better to omalizumab), NOT response per se."
    ),

    "hcp_profiles.influence_network": (
        "CONFIRMED: Optum claims do NOT contain co-authorship, KOL flags, or formal influence graphs. The column influence_network_size (e2i_v3 schema) and peer_influence_score exist but are unpopulated in the Optum converter (convert_optum_rwd.py:1457-1458 set None).\n\n"
        "CLAIMS-DERIVABLE PROXY OPTIONS (build from Optum alone):\n"
        "1) REFERRAL NETWORK (from procedure.parquet if referring_npi is populated):\n"
        "   Directed graph: edge (referring_npi -> rendering_npi) weighted by patient_count.\n"
        "   KOL score = in-degree (specialists who receive many referrals).\n"
        "   Drawback: Optum medical claims often have referring_npi NULL for direct-access specialty visits.\n\n"
        "2) SHARED-PATIENT NETWORK (from medication + procedure):\n"
        "   For each patient, the set of treating HCPs forms a clique. Aggregate to an HCP-HCP edge weighted by shared-patient count.\n"
        "   KOL score = eigenvector centrality of the weighted graph (run with networkx; cache in semantic memory).\n"
        "   Already half-wired: src/memory/semantic_memory.py has get_hcp_influence_network(hcp_id, max_depth=2) and count_hcp_influence_network methods. Backend appears to be FalkorDB (Redis-Graph). Confirm whether the shared-patient or referral edges are being POPULATED in database/memory/ migrations — that wiring may already exist; check before re-implementing.\n\n"
        "3) GEOGRAPHIC-CLINICAL AUTHORITY:\n"
        "   HCPs whose place-of-service codes are predominantly academic medical centers (CMS POS code 22, 21 with academic flag). Provider.parquet's taxonomy isn't sufficient on its own; needs POS + AMC list crosswalk.\n\n"
        "EXTERNAL ENRICHMENT (for canonical KOL data — out of Optum scope):\n"
        "- Definitive Healthcare or HCS Spectrum (commercial KOL DBs)\n"
        "- NIH PubMed + ClinicalTrials.gov for co-authorship and PI roles (we have MCP servers: search_investigators on ClinicalTrials.gov, search_articles on bioRxiv/PubMed — these can be programmatically queried)\n"
        "- CMS Open Payments (transparency.cms.gov) — industry-funded HCP listings; useful as a 'leadership signal' proxy\n"
        "- Doximity API for specialty + network reach (paid, partner-tier API)\n\n"
        "RECOMMENDATION: For Tier 1 Optum-only, populate influence_network_size with SHARED-PATIENT CLIQUE SIZE (cheap, fully derivable). Document as a proxy in data_dictionary.csv. Replace with vendor-grade KOL flags in Tier 2 once external feed is procured."
    ),

    "patient_journeys.source_timestamp": (
        "SCHEMA: source_timestamp, ingestion_timestamp, data_lag_hours are all present on patient_journeys (e2i_v3 schema:859-861). The companion etl_pipeline_metrics table tracks source_data_timestamp + time_to_release_hours at run level (schema:1484-1485) for the TTR KPI.\n\n"
        "OPTUM CDM LIMITATION (correctly identified):\n"
        "Optum parquet drops carry extract_ym at MONTH granularity only (e.g., '202604' for April 2026). There is no per-row source timestamp.\n\n"
        "DERIVATION RULES (conservative, document in methodology_notes):\n"
        "- source_timestamp  = LAST_DAY(extract_ym) at 23:59:59 UTC — worst-case lag assumption; this overstates lag slightly but never understates it.\n"
        "- ingestion_timestamp = file_mtime of the parquet file at receipt (or the parquet's footer metadata timestamp if Optum populates it; check pq.read_metadata(...).created_by)\n"
        "- data_lag_hours    = (ingestion_timestamp - source_timestamp) / 3600\n"
        "Apply at COHORT BUILD TIME, not per-row — every row in the same Optum drop gets the same source_timestamp.\n\n"
        "CLARIFICATION ON FIELD SEMANTICS:\n"
        "Do not confuse vendor delivery timestamps with EVENT timestamps. The latter (fill_dt in medication, admit_date in inpatientdata, fst_dt in lab) describe when the clinical event occurred and are stored on treatment_events. The former (source_timestamp) describes when the data was DELIVERED to us by the vendor.\n\n"
        "ENHANCEMENT REQUESTS TO OPTUM:\n"
        "- Ask vendor for an etl_run_id or data_pull_timestamp key in the parquet metadata. Some Optum panels include this; others don't.\n"
        "- Confirm whether extract_ym refers to data CUTOFF or data DELIVERY date. The KPI definition is sensitive to which one. Without confirmation, default to cutoff = end-of-month."
    ),

    "hcp_intent_surveys.*": (
        "CORRECT — Optum has NO survey data. The hcp_intent_surveys table (e2i_v3 schema:1521-1570) is for SEPARATE market research integration. Designed sources:\n\n"
        "- Periodic ATU studies (Awareness/Trial/Usage) — quarterly market research panels\n"
        "- Conference intercepts at major medical meetings (AAAAI for CSU, ASCO for oncology, ASH for hem)\n"
        "- Veeva CRM detailing follow-up (survey_type='detail_followup')\n"
        "- Sermo, Doximity, M3, ZS Associates HCP panels (proprietary)\n\n"
        "SCHEMA SUPPORTS:\n"
        "- intent_to_prescribe_score (1-7 Likert), intent_to_prescribe_change (delta vs previous survey)\n"
        "- awareness_score, favorability_score (both 1-7)\n"
        "- previous_survey_id, days_since_last_survey, interventions_since_last (JSONB) — enables causal attribution of intent shift to interventions\n"
        "- response_quality_flag for QA exclusion\n\n"
        "INTEGRATION PATH (out of Optum scope):\n"
        "Vendor delivers monthly CSV/Excel survey export -> ingest via the generic FileIngestor (src/agents/ml_foundation/data_preparer/ingestion/) -> map to hcp_intent_surveys.\n\n"
        "INTERIM PROXY (LOW-CONFIDENCE, do NOT use for KPI reporting):\n"
        "Use early-adopter HCPs' rolling prescribing velocity as a behavioral proxy — but this measures ACTION not INTENT, which conflates the two for causal inference. Prefer NULL over proxy here; downstream KPI (Intent-to-Prescribe Delta, BR-002 per src/kpi/calculators/brand_specific.py:136) will simply return null if no surveys exist, and the dashboard should surface the gap explicitly.\n\n"
        "RECOMMENDATION: Leave hcp_intent_surveys empty for the Optum-only Tier 1 deliverable. Do NOT block on this — populate when market research data ships."
    ),

    "user_sessions.*": (
        "CORRECT — user_sessions is platform-generated telemetry (dashboard usage), NOT from claims. Schema at e2i_v3:1326-1349.\n\n"
        "POPULATION REQUIRES FRONTEND INSTRUMENTATION (frontend/ React+TS):\n"
        "- Session start/end events -> POST /api/sessions on auth and tab close\n"
        "- page_views, queries_executed, triggers_viewed, actions_taken, exports_downloaded -> JS click handlers and route guards\n"
        "- device_type, browser -> User-Agent parsing in middleware\n"
        "- engagement_score -> nightly derived as (page_views + 2*queries + 3*actions) / session_duration_minutes, normalized 0-9.99\n\n"
        "KPI SUPPORT: The MAU/WAU/DAU KPIs (v_kpi_active_users view in docs/data/02-CORE-DATA-DICTIONARY.md:1679) depend on this table. Without instrumentation, those KPIs return zero — which is honest (no active users measured) but operationally unhelpful.\n\n"
        "OUT OF SCOPE for the Optum data conversion task. Add to a separate frontend-instrumentation epic; coordinate with the React/TypeScript team in frontend/."
    ),

    "causal_paths.*": (
        "CORRECT — causal_paths is OUTPUT of the causal_impact agent (Tier 2), NOT derivable from source data. Schema at e2i_v3:1267-1316.\n\n"
        "AGENT PIPELINE: src/agents/causal_impact/agent.py — pipeline order: graph_builder -> estimation -> refutation -> sensitivity -> interpretation. SLA: 120s total. Methods supported (per `method_used` enum in the data dictionary): dowhy_refutation, pc_algorithm, granger, transfer_entropy.\n\n"
        "DEPENDENCY: causal_impact consumes patient_journeys + treatment_events as INPUT and produces causal_paths as OUTPUT. So the Optum data conversion is the UPSTREAM dependency.\n\n"
        "LEAKAGE-SAFETY: The Optum cohort's data_split column flows through ml_split_registry, which is the same registry the causal_impact agent uses. By construction, training vs validation isolation is preserved end-to-end. No additional action needed; just confirm the agent's queries filter on data_split='train' for fit and data_split='test' for evaluation.\n\n"
        "ACTION: Once the Optum cohort is converted and loaded, trigger causal_impact agent against it and confirm causal_paths is populated. Track via agent_activities table (e2i_v3 schema:1149)."
    ),

    "ml_predictions.*": (
        "CORRECT — ml_predictions is OUTPUT of model inference (prediction_synthesizer Tier 4 agent), NOT derivable from source data. Schema at e2i_v3:992-1064.\n\n"
        "AGENT: src/agents/prediction_synthesizer/agent.py — ensembles multiple base models (xgboost, lightgbm, causal_forest per `model_type` enum). Ensemble methods: average, weighted, stacking, voting.\n\n"
        "WRITES (40+ columns): prediction_value, confidence_score, shap_values (per-prediction), top_features, model_auc/precision/recall, calibration_score, model_pr_auc (WS1 gap), rank_metrics (recall@K), brier_score, treatment_effect_estimate (ATE), heterogeneous_effect (CATE), segment_assignment, causal_confidence, counterfactual_outcome.\n\n"
        "GROUND-TRUTH LOOP — important for the Optum cohort:\n"
        "actual_outcome / outcome_recorded_at / outcome_source / outcome_observation_window_days are the columns the converter's targets (initiated_biologic_180d / discontinued_180d / persistent_at_180d, per convert_optum_rwd.py:1201-1205) feed back into for drift detection. This is the closed feedback loop driven by database/migrations/006_feedback_loop_infrastructure.sql.\n\n"
        "ACTION: Confirm the Optum cohort's test-split rows flow into the feedback_loop tasks (src/tasks/feedback_loop_tasks.py — already wired for market_share_impact, risk, treatment_response). For each prediction written, the corresponding ground-truth outcome should be backfilled after the prediction observation window elapses."
    ),

    "patient_journeys.payer_category": (
        "DERIVATION RULES (Optum CDM source fields):\n"
        "- demographics.bus       -> {COM: commercial, MCR: medicare, MCD: medicaid, OTH: other}\n"
        "- demographics.product   -> plan_type granularity (PPO/HMO/POS/HDHP/EPO for COM; MAPD/PDP/SUPP for MCR)\n"
        "- demographics.health_exch (BOOLEAN) -> ACA marketplace flag (TRUE & bus=COM => payer_category='commercial_exchange')\n"
        "- demographics.lis_dual  (BOOLEAN) -> low-income subsidy / dual-eligible flag (TRUE & bus=MCR => payer_category='medicare_lis_dual')\n\n"
        "PROPOSED payer_category ENUM (extending the current insurance_type VARCHAR(20)):\n"
        "  commercial, commercial_exchange, medicare, medicare_advantage, medicare_lis_dual, medicaid, cash, other\n\n"
        "CURRENT CODEBASE STATE:\n"
        "- scripts/rwd_common.py:43-47 has only the basic 3-way INSURANCE_TYPE_MAP (COM/MCR/MCD). This needs extension to capture lis_dual + health_exch.\n"
        "- convert_optum_rwd.py:801 calls rwdc.insurance_type(demo_row.get('bus')) -- single field; does not consume lis_dual or health_exch yet.\n"
        "- patient_journeys.insurance_type is VARCHAR(20) and does not yet have a payer_category companion column. Schema migration required (or repurpose insurance_type with widened vocabulary).\n"
        "- TRACKED: GitHub issue #156 (PR B) item 6 covers the bus/product/health_exch/lis_dual derivation and the payer_category column migration.\n\n"
        "SPECIALTY PHARMACY CHANNEL — TRACTABLE VIA NPPES (vendor has confirmed NPI preservation in our extract):\n"
        "The earlier assumption that Optum drops the dispenser NPI was wrong — vendor confirmed dispenser/prescriber/rendering/facility NPIs are all preserved. NPPES (CMS National Plan and Provider Enumeration System) is a free public API + monthly bulk dump that gives us NUCC taxonomy codes per NPI.\n"
        "APPROACH:\n"
        "(a) Ingest the monthly NPPES bulk dump (~10 GB) into a local npi_taxonomy table; refresh monthly via Celery beat.\n"
        "(b) Resolve each medication.dispenser_npi against the cache; match NUCC taxonomy codes 3336S0011X (Specialty Pharmacy), 3336M0002X (Mail Order Pharmacy — frequently specialty), 3336H0001X (Home Infusion Therapy Pharmacy — frequently specialty). Set specialty_pharmacy_flag = TRUE on any match.\n"
        "(c) API fallback for cache-miss NPIs: https://npiregistry.cms.hhs.gov/api/?number=<NPI>&version=2.1 (free, no auth, rate-limited — fine for new-NPI deltas).\n"
        "TRACKED: GitHub issue #154 (PR B-prime: NPPES NPI taxonomy enrichment) carries the full scope — specialty_pharmacy_flag derivation lands as item 5 of that PR.\n"
        "BONUS: the same NPPES integration unlocks ~8 currently-None fields on hcp_profiles (specialty, sub_specialty, practice_type, practice_size, years_experience, affiliation_primary, state/city/zip_code, geographic_region, academic_hcp) and sharpens §7.7 provider-mix features. Foundational ETL capability, not a one-off pharmacy flag lookup.\n\n"
        "BEST PRACTICE — STORE GRANULAR + NORMALIZED:\n"
        "Persist BOTH the raw source fields (bus, product, health_exch, lis_dual) AS-IS on patient_journeys AND the derived normalized payer_category. This enables future re-derivation without re-ETL and preserves audit trail. Reference the v3 schema's source_combination_method column (e2i_v3:858) as the existing precedent for keeping raw + derived together."
    ),
}


# Visual styling — match openpyxl's auto-render to keep the file readable.
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

BODY_ALIGN = Alignment(wrap_text=True, vertical="top", horizontal="left")
BODY_FONT = Font(name="Calibri", size=10)


def main() -> None:
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, DST)

    wb = openpyxl.load_workbook(DST)
    ws = wb["Gap features"]

    # Label the orphan column F (only F7 had pre-existing content for engagement_score weighting).
    if ws["F1"].value is None:
        ws["F1"] = "Additional Comments"

    # New column G: Claude Code Analysis & Recommendations.
    ws["G1"] = "Claude Code Analysis & Recommendations"

    # Style row-1 header for the two columns we touched.
    for cell in (ws["F1"], ws["G1"]):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Re-style row 1 fully to keep it consistent.
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        if cell.value is None:
            continue
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Fill column G per row using A as the E2I-field key.
    missing: list[str] = []
    for row in range(2, ws.max_row + 1):
        e2i_field = ws.cell(row=row, column=1).value
        if not e2i_field:
            continue
        key = str(e2i_field).strip()
        text = RESPONSES.get(key)
        if text is None:
            missing.append(key)
            continue
        cell = ws.cell(row=row, column=7, value=text)
        cell.alignment = BODY_ALIGN
        cell.font = BODY_FONT

    if missing:
        raise SystemExit(f"Missing response keys for: {missing}")

    # Apply wrap_text to existing populated cells so the new wide column doesn't ruin layout.
    for row in range(2, ws.max_row + 1):
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            if c.value is not None and c.alignment.wrap_text is not True:
                c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # Column widths tuned for readability: A narrow (field name), B-F medium, G wide.
    widths = {"A": 32, "B": 22, "C": 38, "D": 50, "E": 50, "F": 50, "G": 90}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights: header tall, body adaptive (openpyxl won't auto-size, so set a generous default).
    ws.row_dimensions[1].height = 34
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 260

    # Freeze the header + first column so the wide G column is comfortable to read.
    ws.freeze_panes = "B2"

    wb.save(DST)
    print(f"Wrote {DST}")


if __name__ == "__main__":
    main()
